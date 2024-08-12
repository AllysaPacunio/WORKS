import openpyxl
from openpyxl import Workbook, load_workbook
import clips
import tkinter as tk
from tkinter import messagebox,ttk
from ttkbootstrap import Style

excel = load_workbook('RIASEC_EXAM.xlsx')
sheet1= excel["QUESTIONS"]
sheet2=excel["OPTIONS"]

questionnaire_data=[]
answers= []

for questions in sheet1.iter_rows( min_row=2, max_row=43, min_col=1, max_col=1, values_only=True):
    questionnaire= {}
    questionnaire["question"]= questions
    options= ['Strongly_Agree', 'Agree', 'Neutral', 'Disagree', 'Strongly_Disagree']
    questionnaire["options"] = options
    questionnaire_data.append(questionnaire)
    
print(questionnaire_data)

def show_question():
    qst= questionnaire_data[current_question]
    questions_label.config(text=qst["question"])

    options = qst["options"]
    for i in range(5):
        choice_buttons[i].config(text=options[i], state="normal")
    
    feedback_label.config(text="")
    next_button.config(state="disabled")

def answer(choice):
    selected_choice = str(choice_buttons[choice].cget("text"))
    
    env = clips.Environment()
    env.build('(deftemplate my-template (multislot numbers))')
    my_fact = env.assert_string('(my-template (numbers))')
    my_fact['numbers'] = clips.Multislot([1, 2, 3, 4, 5])
    print(my_fact)
   
    
    

# (printout t "Hello, {selected_choice}" crlf))
    







    global current_question
    current_question +=1
    score_label.config(text= "{}/{} questions".format(current_question, len(questionnaire_data)))
    feedback_label.config(text= "{}/{} questions".format(current_question, len(questionnaire_data)), foreground="green")

    for button in choice_buttons:
        button.config(state="disabled")
        next_button.config(state="normal")



def next_question():
    if current_question < len(questionnaire_data):
        show_question()

    else:
        messagebox.showinfo("Quiz Completed",
                            "Quiz Completed! Final score: {}/{}".format(current_question, len(questionnaire_data)))
        root.destroy()


root = tk.Tk()
root.title("Career Assessment App")
root.geometry("600x500")
style = Style(theme="flatly") 

questions_label = ttk.Label(root, anchor="center", wraplength=500, padding= 10)
questions_label.pack(pady=10)

choice_buttons= []
for i in range(5):
    button = ttk.Button(root, command=lambda i=i: answer(i))

    button.pack(pady=4)
    choice_buttons.append(button)

feedback_label = ttk.Label(root, anchor= "center", padding= 10)
feedback_label.pack(pady=10)

# total_answered = 0 


score_label = ttk.Label(root, text="Score 0/{}".format(len(questionnaire_data)), anchor="center", padding=10)

next_button = ttk.Button(root, text="Next", command=next_question, state="disabled")
next_button.pack(pady=10)

current_question = 0

show_question()


root.mainloop()

# import clips
# import PySimpleGUI as sg


# RULES = [
#     """
#     (defrule book-service
#       =>
#       (bind ?answer (polar-question "Are you a first-time user?"))
#       (assert (first-time-user ?answer)))
#     """,
#     """
#     (defrule first-timer
#       (or first-time-user "Yes" "No")
#       =>
#       (bind ?answer (polar-question "Do you like reading books?"))
#       (assert (likes-reading-books ?answer)))
#     """
# ]

# main = Tk()
# def display_title(self):
#     """To display title"""

# title = Label(self.window, text="iQuiz Application",
#                       width=50, bg="green", fg="white", font=("ariel", 20, "bold"))
#     title.place(x=0, y=2)
    


# def main():
#     env = clips.Environment()
#     env.define_function(polar_question, name='polar-question')
#     for rule in RULES:
#         env.build(rule)
#     env.run()



# root = Tk() 
# root.title("Career Assesment Exam")
# root.geometry("850x530")

# title = Label(root, text="Career Assesment Exam", width=50, bg="pink", fg="black", font=("ariel", 20, "bold"))
# title.place(x=0, y=2)

# instruction = Label(root, text="Be honest in answering the following question for accurate result.", width=50, fg="black", font=("ariel", 10, "bold"))
# instruction.place(x=0, y=50)

# root.mainloop() 

